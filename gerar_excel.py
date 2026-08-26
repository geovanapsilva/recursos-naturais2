import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import ScatterChart, LineChart, Reference
import numpy as np

# Criar novo workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove a planilha padrão

# ========== PLANILHA 1: EQUILÍBRIO DE ESTADO ESTACIONÁRIO ==========
ws1 = wb.create_sheet("Planilha 1")

# Configurar parâmetros
ws1['A1'] = "Planilha 1"
ws1['A3'] = "r"
ws1['A4'] = "K"
ws1['A5'] = "a"

ws1['B3'] = 1
ws1['B4'] = 1
ws1['B5'] = 0.5

# Formatar células amarelas
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ws1['B3'].fill = yellow_fill
ws1['B4'].fill = yellow_fill
ws1['B5'].fill = yellow_fill

# Adicionar rótulos
ws1['A6'] = "Xss"
ws1['A7'] = "Yss"
ws1['A8'] = "|G'(Xss)|"
ws1['A9'] = "Xo"

# Fórmulas
ws1['B6'] = "=($B$4*($B$3-$B$5))/$B$3"  # Xss
ws1['B7'] = "=$B$3*$B$6*(1-$B$6/$B$4)"  # Yss
ws1['B8'] = "=ABS(1-$B$3+$B$5)"  # |G'(Xss)|
ws1['B9'] = 0.1  # Xo

# Tabela de valores (X, F(X), Y)
ws1['A12'] = "X"
ws1['B12'] = "F(X) = r*X*(1-X/K)"
ws1['C12'] = "Y = a*X"

# Preencher coluna X (0 a 1 em passos de 0.025)
for i in range(41):  # 0 a 1 em passos de 0.025
    row = 13 + i
    ws1[f'A{row}'] = i * 0.025

# Fórmulas para F(X) e Y
for i in range(41):
    row = 13 + i
    if i == 0:
        ws1[f'B{row}'] = f"=A{row}"
        ws1[f'C{row}'] = f"=A{row}"
    else:
        ws1[f'B{row}'] = f"=$B$3*A{row}*(1-A{row}/$B$4)"
        ws1[f'C{row}'] = f"=$B$5*A{row}"

# Criar gráfico para Planilha 1
chart1 = ScatterChart()
chart1.title = 'Equilíbrio no "estado estacionário"'
chart1.x_axis.title = 'X'
chart1.y_axis.title = 'F(X) e Y'

xvalues1 = Reference(ws1, min_col=1, min_row=13, max_row=53)
yvalues_fx = Reference(ws1, min_col=2, min_row=12, max_row=53)
yvalues_y = Reference(ws1, min_col=3, min_row=12, max_row=53)

chart1.add_data(yvalues_fx, titles_from_data=True)
chart1.add_data(yvalues_y, titles_from_data=True)
chart1.set_categories(xvalues1)

ws1.add_chart(chart1, "E2")

# ========== PLANILHA 2: CAMINHO ÓTIMO ==========
ws2 = wb.create_sheet("Planilha 2")

# Copiar parâmetros
ws2['A3'] = "r"
ws2['A4'] = "K"
ws2['A5'] = "a"

ws2['B3'] = 1
ws2['B4'] = 1
ws2['B5'] = 0.5

ws2['B3'].fill = yellow_fill
ws2['B4'].fill = yellow_fill
ws2['B5'].fill = yellow_fill

ws2['A6'] = "Xss"
ws2['A7'] = "Yss"
ws2['A8'] = "|G'(Xss)|"
ws2['A9'] = "Xo"

ws2['B6'] = "=($B$4*($B$3-$B$5))/$B$3"
ws2['B7'] = "=$B$3*$B$6*(1-$B$6/$B$4)"
ws2['B8'] = "=ABS(1-$B$3+$B$5)"
ws2['B9'] = 0.1

# Tabela dinâmica
ws2['A12'] = "t"
ws2['B12'] = "Xt"
ws2['C12'] = "Yt"

for i in range(20):
    ws2[f'A{13+i}'] = i

ws2['B13'] = "=$B$9"  # X0

for i in range(1, 20):
    row = 13 + i
    ws2[f'B{row}'] = f"=B{row-1}*(1+$B$3-$B$5-$B$3*B{row-1}/$B$4)"

for i in range(20):
    row = 13 + i
    ws2[f'C{row}'] = f"=$B$5*B{row}"

# Criar gráfico para Planilha 2
chart2 = ScatterChart()
chart2.title = '"Caminhos ótimos" para Xt e Yt'
chart2.x_axis.title = 't'
chart2.y_axis.title = 'Xt e Yt'

xvalues2 = Reference(ws2, min_col=1, min_row=13, max_row=32)
yvalues_xt = Reference(ws2, min_col=2, min_row=12, max_row=32)
yvalues_yt = Reference(ws2, min_col=3, min_row=12, max_row=32)

chart2.add_data(yvalues_xt, titles_from_data=True)
chart2.add_data(yvalues_yt, titles_from_data=True)
chart2.set_categories(xvalues2)

ws2.add_chart(chart2, "E2")

# ========== PLANILHA 3: POLÍTICA DE EXTRAÇÃO LINEAR ÓTIMA ==========
ws3 = wb.create_sheet("Planilha 3")

# Parâmetros principais
ws3['A3'] = "r"
ws3['A4'] = "K"
ws3['A5'] = "a"

ws3['B3'] = 1
ws3['B4'] = 1
ws3['B5'] = 0.524881  # Valor otimizado

ws3['B3'].fill = yellow_fill
ws3['B4'].fill = yellow_fill

blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
ws3['B5'].fill = blue_fill
ws3['B5'].font = Font(color="FFFFFF")

ws3['A6'] = "Xss"
ws3['A7'] = "Yss"
ws3['A8'] = "|G'(Xss)|"
ws3['A9'] = "Xo"

ws3['B6'] = "=($B$4*($B$3-$B$5))/$B$3"
ws3['B7'] = "=$B$3*$B$6*(1-$B$6/$B$4)"
ws3['B8'] = "=ABS(1-$B$3+$B$5)"
ws3['B9'] = 0.1

# Parâmetros econômicos
ws3['A11'] = "p"
ws3['B11'] = 5
ws3['A12'] = "c"
ws3['B12'] = 1
ws3['A13'] = "δ"
ws3['B13'] = 0.05
ws3['A14'] = "ρ"
ws3['B14'] = "=1/(1+B13)"

# Dinâmica
ws3['A16'] = "t"
ws3['B16'] = "Xt"
ws3['C16'] = "Yt"
ws3['D16'] = "π(t)"
ws3['E16'] = "ρᵗ*π(t)"

for i in range(20):
    ws3[f'A{17+i}'] = i

ws3['B17'] = "=$B$9"

for i in range(1, 20):
    row = 17 + i
    ws3[f'B{row}'] = f"=B{row-1}*(1+$B$3-$B$5-$B$3*B{row-1}/$B$4)"

for i in range(20):
    row = 17 + i
    ws3[f'C{row}'] = f"=$B$5*B{row}"
    ws3[f'D{row}'] = f"=($B$11-$B$12)*C{row}"
    ws3[f'E{row}'] = f"=$B$14^A{row}*D{row}"

# Célula de objetivo
ws3['E37'] = "=SUM(E17:E36)"
ws3['E37'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
ws3['E37'].font = Font(bold=True, color="FFFFFF")

# Criar gráfico para Planilha 3
chart3 = ScatterChart()
chart3.title = '"Caminhos ótimos" para Xt e Yt'
chart3.x_axis.title = 't'
chart3.y_axis.title = 'Xt e Yt'

xvalues3 = Reference(ws3, min_col=1, min_row=17, max_row=36)
yvalues3_xt = Reference(ws3, min_col=2, min_row=16, max_row=36)
yvalues3_yt = Reference(ws3, min_col=3, min_row=16, max_row=36)

chart3.add_data(yvalues3_xt, titles_from_data=True)
chart3.add_data(yvalues3_yt, titles_from_data=True)
chart3.set_categories(xvalues3)

ws3.add_chart(chart3, "G2")

# ========== PLANILHA 4: SIMULAÇÕES ==========
ws4 = wb.create_sheet("Planilha 4 - Simulações")

ws4['A1'] = "Planilha 4 - Simulações com diferentes parâmetros"
ws4['A3'] = "Faça simulações alterando os parâmetros (r, K, a, X₀) nas Planilhas 1 e 2"
ws4['A5'] = "Exemplos de simulações:"
ws4['A6'] = "1. Mude 'a' para 1.3 ou 2.6"
ws4['A7'] = "2. Mude 'r' para 2.0 ou 0.5"
ws4['A8'] = "3. Observe como os gráficos das Planilhas 1 e 2 mudam"

# Salvar arquivo
caminho_arquivo = "Atividade_Economia_Completa.xlsx"
wb.save(caminho_arquivo)

print(f"✅ Arquivo '{caminho_arquivo}' criado com sucesso!")
print("📊 Contém 4 planilhas com todas as fórmulas e gráficos prontos!")
print("🎯 Abra no Excel e use normalmente!")
